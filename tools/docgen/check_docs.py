"""Docs & traceability consistency validator for DigitalSecretary.

Fails (exit 1) on drift between artifacts. Wired into build.ps1.
Checks:
  1. Spreadsheets match their source data tables (not hand-edited / not stale).
  2. Every path reference in the traceability matrix exists on disk.
  3. Referenced code/test symbols (method/test names) appear in their files (warning).
  4. Requirement IDs are consistent: referenced IDs exist; every functional requirement is traced;
     each requirement's ID appears in its requirement doc.
  5. Traceability Coverage has no Gap (all required artifact cells present).
  6. Every src/Features/<X> has a full artifact set (FEATURE.md, requirement doc, user-guide, a row).
  7. Generated text (user manual + spreadsheets) is clean (no mojibake, no unresolved tokens).

Run:  python tools/docgen/check_docs.py
"""
import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

import build_excel_docs as src

ROOT = Path(__file__).resolve().parents[2]
REQ_XLSX = ROOT / "docs/requirements/DigitalSecretary-Requirements.xlsx"
TRACE_XLSX = ROOT / "docs/traceability/DigitalSecretary-Traceability-Matrix.xlsx"
MANUAL = ROOT / "docs/user-guide/DigitalSecretary-User-Manual.html"

ID_RE = re.compile(r"\b(?:APP|NFR|LAU|CAL|CLP|EML)-\d+\b")
MOJIBAKE = ["Â", "Ã", "â"]
ARTIFACT_COLS = {4: "Requirement Doc", 5: "Mock/Screen", 6: "Code", 7: "Unit Test",
                 8: "QA", 9: "User Manual", 10: "Architecture", 11: "Code Doc"}
REQUIRED_COVERAGE_COLS = (4, 6, 7, 8, 9, 11)  # Req Doc, Code, Unit Test, QA, User Manual, Code Doc

errors, warnings = [], []
def err(m): errors.append(m)
def warn(m): warnings.append(m)
def norm(v): return ("" if v is None else str(v)).strip()


def paths_in(cell):
    s = norm(cell)
    if not s or s == "-" or s.lower().startswith("n/a"):
        return []
    s = re.sub(r"\([^)]*\)", "", s)  # drop parentheticals (method/test names)
    out = []
    for tok in s.split(","):
        tok = tok.strip().split("#")[0].strip()
        if "/" in tok:
            out.append(tok)
    return out


def symbols_in(cell):
    return [m.strip() for grp in re.findall(r"\(([^)]*)\)", norm(cell))
            for m in grp.split(",") if m.strip()]


def main():
    for f in (REQ_XLSX, TRACE_XLSX, MANUAL):
        if not f.exists():
            err(f"Missing generated artifact: {f.relative_to(ROOT)} (run the generators).")
    if errors:
        return report()

    req_ws = load_workbook(REQ_XLSX).worksheets[0]
    trace_wb = load_workbook(TRACE_XLSX)
    trace_ws = trace_wb["Traceability Matrix"]
    manual_text = MANUAL.read_text(encoding="utf-8", errors="ignore")

    # 1. Staleness vs source data tables
    if req_ws.max_row - 1 != len(src.REQUIREMENTS):
        err(f"Requirements.xlsx has {req_ws.max_row - 1} rows; source has {len(src.REQUIREMENTS)} (stale - regenerate).")
    else:
        for i, exp in enumerate(src.REQUIREMENTS):
            for c in range(1, len(src.REQ_HEADERS) + 1):
                if norm(req_ws.cell(i + 2, c).value) != norm(exp[c - 1]):
                    err(f"Requirements.xlsx stale at row {i + 2}, col {c} (regenerate).")
                    break
    if trace_ws.max_row - 1 != len(src.TRACE):
        err(f"Traceability.xlsx has {trace_ws.max_row - 1} rows; source has {len(src.TRACE)} (stale - regenerate).")
    else:
        for i, exp in enumerate(src.TRACE):
            stale = any(norm(trace_ws.cell(i + 2, c).value) != norm(exp[c - 1]) for c in range(1, 12))
            stale = stale or norm(trace_ws.cell(i + 2, 13).value) != norm(exp[11])
            if stale:
                err(f"Traceability.xlsx stale at row {i + 2} (regenerate).")

    # 2/3/5. Path existence, symbols, coverage (per traceability row)
    for r in range(2, trace_ws.max_row + 1):
        feat, sub = norm(trace_ws.cell(r, 1).value), norm(trace_ws.cell(r, 2).value)
        tag = f"[{feat} / {sub}]"
        for c, label in ARTIFACT_COLS.items():
            for p in paths_in(trace_ws.cell(r, c).value):
                if not (ROOT / p).exists():
                    err(f"{tag} {label} path not found: {p}")
        anchor = re.search(r"#([\w\-]+)", norm(trace_ws.cell(r, 9).value))
        if anchor and f'id="{anchor.group(1)}"' not in manual_text:
            err(f"{tag} user-manual anchor #{anchor.group(1)} not found in the manual.")
        for c in (6, 7, 8):  # code / unit / qa symbol presence
            cell = trace_ws.cell(r, c).value
            files = [ROOT / p for p in paths_in(cell) if (ROOT / p).is_file()]
            text = "".join(f.read_text(encoding="utf-8", errors="ignore") for f in files)
            for sym in symbols_in(cell):
                token = sym.split()[0]
                if files and re.fullmatch(r"\w+", token) and token not in text:
                    warn(f"{tag} symbol '{token}' not found in {', '.join(p.name for p in files)}")
        missing = [ARTIFACT_COLS[c] for c in REQUIRED_COVERAGE_COLS if not norm(trace_ws.cell(r, c).value)]
        if missing:
            err(f"{tag} Coverage GAP - missing {missing}")

    # 4. Requirement ID consistency
    req_ids = {}
    for r in range(2, req_ws.max_row + 1):
        rid, itype, doc = (norm(req_ws.cell(r, 1).value), norm(req_ws.cell(r, 2).value),
                           norm(req_ws.cell(r, 11).value))
        if not rid:
            continue
        req_ids[rid] = itype
        dp = ROOT / doc
        if not dp.is_file():
            err(f"Requirement {rid}: doc missing {doc}")
        elif rid not in dp.read_text(encoding="utf-8", errors="ignore"):
            err(f"Requirement {rid} not found in its doc {doc}")
    traced = set()
    for r in range(2, trace_ws.max_row + 1):
        for rid in ID_RE.findall(norm(trace_ws.cell(r, 3).value)):
            traced.add(rid)
            if rid not in req_ids:
                err(f"Traceability row {r} references unknown requirement {rid}")
    for rid, itype in req_ids.items():
        if itype != "NFR" and rid not in traced:
            err(f"Functional requirement {rid} is not covered by any traceability row")

    # 6. Feature completeness
    for d in sorted((ROOT / "src/Features").iterdir()):
        if not d.is_dir():
            continue
        pj = d / "plugin.json"
        if not pj.is_file():
            err(f"Feature folder {d.name} has no plugin.json")
            continue
        fid = json.loads(pj.read_text(encoding="utf-8")).get("id", "")
        for label, p in {"FEATURE.md": d / "FEATURE.md",
                          "requirement doc": ROOT / f"docs/requirements/features/{fid}.md",
                          "user-guide doc": ROOT / f"docs/user-guide/features/{fid}.md"}.items():
            if not p.is_file():
                err(f"Feature '{fid}' missing {label}: {p.relative_to(ROOT)}")
        if not any(f"src/Features/{d.name}/" in norm(trace_ws.cell(r, 6).value)
                   for r in range(2, trace_ws.max_row + 1)):
            err(f"Feature '{fid}' has no traceability row referencing src/Features/{d.name}/")

    # 7. Cleanliness of generated text
    for ch in MOJIBAKE:
        if ch in manual_text:
            err(f"User manual contains encoding artifact U+{ord(ch):04X}")
    if re.search(r"\{\{[a-z0-9\-]+\}\}", manual_text):
        err("User manual has unresolved {{tokens}}")
    for ws in trace_wb.worksheets + load_workbook(REQ_XLSX).worksheets:
        for row in ws.iter_rows(values_only=True):
            for v in row:
                if isinstance(v, str) and any(ch in v for ch in MOJIBAKE):
                    err(f"Spreadsheet cell has encoding artifact: {v[:40]!r}")

    return report()


def report():
    print("=== Docs & traceability consistency ===")
    for e in errors:
        print("  ERROR:", e)
    for w in warnings:
        print("  WARN :", w)
    if not errors and not warnings:
        print("  All artifacts consistent.")
    print(f"RESULT errors={len(errors)} warnings={len(warnings)}")
    return 1 if errors else 0


if __name__ == "__main__":
    sys.exit(main())
